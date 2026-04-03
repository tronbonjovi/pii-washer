import io

from pii_washer.extractors.base import BaseExtractor


class XlsxExtractor(BaseExtractor):
    """Extract plain text from .xlsx files as pipe-delimited rows.

    Each row is converted to pipe-delimited format:
    "| cell1 | cell2 | ... |"

    Multiple sheets are each prefixed with "[SheetName]" and separated
    by blank lines. Empty rows are skipped.
    """

    def extract(self, content: bytes, filename: str) -> str:
        """Extract structured plain text from .xlsx bytes.

        Args:
            content: Raw bytes of the .xlsx file.
            filename: Original filename, used in error messages.

        Returns:
            Pipe-delimited rows. Multiple sheets are labelled with
            "[SheetName]" prefixes and separated by blank lines.

        Raises:
            RuntimeError: If openpyxl is not installed.
            ValueError: If the file cannot be parsed or contains no text.
        """
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required to read .xlsx files. "
                "Install it with: pip install openpyxl"
            ) from exc

        try:
            wb = load_workbook(
                io.BytesIO(content), read_only=True, data_only=True
            )
        except Exception as exc:
            raise ValueError(
                f"'{filename}' could not be read as an .xlsx file: {exc}"
            ) from exc

        try:
            sheet_blocks: list[str] = []
            multiple_sheets = len(wb.sheetnames) > 1

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_output: list[str] = []

                for row in ws.iter_rows():
                    cells = [
                        str(cell.value) if cell.value is not None else ""
                        for cell in row
                    ]
                    # Skip rows where all cells are empty
                    if not any(c.strip() for c in cells):
                        continue
                    rows_output.append("| " + " | ".join(cells) + " |")

                if not rows_output:
                    continue

                rows_text = "\n".join(rows_output)
                block = f"[{sheet_name}]\n\n{rows_text}" if multiple_sheets else rows_text

                sheet_blocks.append(block)

        finally:
            wb.close()

        if not sheet_blocks:
            raise ValueError(
                f"No text content found in '{filename}'."
            )

        return "\n\n".join(sheet_blocks)
