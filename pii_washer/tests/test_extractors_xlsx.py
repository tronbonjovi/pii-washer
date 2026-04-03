import io

import pytest
from openpyxl import Workbook

from pii_washer.extractors.xlsx import XlsxExtractor


# === Helpers ===


def _make_xlsx(sheets: dict[str, list[list]]) -> bytes:
    """Build an .xlsx file in memory.

    Args:
        sheets: dict mapping sheet name to a list of rows,
                each row a list of cell values.

    Returns:
        Raw bytes of the .xlsx file.
    """
    wb = Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def extractor():
    return XlsxExtractor()


# === Happy path ===


def test_returns_string_type(extractor):
    content = _make_xlsx({"Sheet1": [["Name", "Email"], ["John Smith", "john@example.com"]]})
    result = extractor.extract(content, "test.xlsx")
    assert isinstance(result, str)


def test_single_sheet_extraction(extractor):
    content = _make_xlsx({"Sheet1": [["Name", "Email"], ["John Smith", "john@example.com"]]})
    result = extractor.extract(content, "test.xlsx")
    assert "| Name | Email |" in result
    assert "| John Smith | john@example.com |" in result


def test_single_sheet_no_label(extractor):
    """Single-sheet workbooks should not include a [SheetName] label."""
    content = _make_xlsx({"Data": [["A", "B"], ["1", "2"]]})
    result = extractor.extract(content, "test.xlsx")
    assert "[Data]" not in result


def test_multi_sheet_labelled(extractor):
    """Multi-sheet workbooks prefix each section with [SheetName]."""
    content = _make_xlsx({
        "Employees": [["Name", "Role"], ["Alice", "Engineer"]],
        "Contacts": [["Phone"], ["555-1234"]],
    })
    result = extractor.extract(content, "test.xlsx")
    assert "[Employees]" in result
    assert "[Contacts]" in result
    assert "| Alice | Engineer |" in result
    assert "| 555-1234 |" in result


def test_multi_sheet_sections_separated(extractor):
    """Sections for different sheets are separated by blank lines."""
    content = _make_xlsx({
        "Sheet1": [["A"]],
        "Sheet2": [["B"]],
    })
    result = extractor.extract(content, "test.xlsx")
    assert "\n\n" in result


def test_numeric_cell_values(extractor):
    """Numeric cell values are converted to strings."""
    content = _make_xlsx({"Sheet1": [["Name", "Age"], ["Bob", 42]]})
    result = extractor.extract(content, "test.xlsx")
    assert "42" in result


def test_empty_rows_skipped(extractor):
    """Rows where all cells are empty are skipped."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Email"])
    ws.append([None, None])
    ws.append(["Alice", "alice@example.com"])
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    result = extractor.extract(content, "test.xlsx")
    lines = [l for l in result.splitlines() if l]
    assert len(lines) == 2


# === Error handling ===


def test_empty_workbook_raises_value_error(extractor):
    """A workbook with no content in any sheet raises ValueError."""
    wb = Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    with pytest.raises(ValueError, match="No text content"):
        extractor.extract(content, "empty.xlsx")


def test_corrupted_file_raises_value_error(extractor):
    garbage = b"This is not a valid xlsx file \x00\x01\x02\x03"
    with pytest.raises(ValueError, match="could not be read"):
        extractor.extract(garbage, "bad.xlsx")
